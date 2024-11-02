import openpyxl

book = openpyxl.load_workbook("C:/Users/Usuario/Documents/Libro1.xlsx")
sheet = book.active
#to retrieve a value from a specific cell
dict = {}

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":

        for j in range(2,sheet.max_column):
            dict[sheet.cell(row=1, column=j).value] = (sheet.cell(row=i, column=j).value)

print(dict)