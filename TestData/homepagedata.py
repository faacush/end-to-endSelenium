import openpyxl


class HomePageData:

    test_HomePage_data=[{"firstname": "Rahul", "lastname": "Shetty", "gender": "Male"},{"firstname": "Facundo", "lastname": "Navarro", "gender": "Female"}]

    @staticmethod
    def getData(test_case_name):
        dict = {}
        book = openpyxl.load_workbook("C:/Users/Usuario/Documents/Libro1.xlsx")
        sheet = book.active

        for i in range(1,sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2,sheet.max_column):
                    dict[sheet.cell(row=1, column=j).value] = (sheet.cell(row=i, column=j).value)

        return[dict]
